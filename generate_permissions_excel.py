import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def create_permissions_excel(data_text, output_file='permisos.xlsx'):
    # Procesar el texto de entrada
    lines = data_text.strip().split('\n')
    
    # Preparar los datos para el DataFrame
    data = []
    current_object = None
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        if '\t' not in line:
            current_object = line
        else:
            permissions = line.split('\t')
            permissions = [p.strip() for p in permissions]
            # Ahora guardamos True/False en lugar de símbolos
            data.append([current_object] + [
                True if p.lower() == 'checked' else 
                False if p.lower() == 'not checked' else 
                False for p in permissions
            ])

    # Crear DataFrame
    columns = ['Object Name', 'Create', 'Read', 'Edit', 'Delete', 
              'View All\nRecords', 'Modify All\nRecords', 'View All\nFields']
    df = pd.DataFrame(data, columns=columns)
    
    # Ordenar el DataFrame alfabéticamente por Object Name
    df = df.sort_values('Object Name', ignore_index=True)
    
    # Guardar en Excel
    df.to_excel(output_file, index=False, sheet_name='Permissions')
    
    # Dar formato al archivo Excel
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    
    # Estilos
    header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    centered = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Aplicar estilos al encabezado
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = centered
        cell.border = border

    # Ajustar el ancho de las columnas y aplicar estilos a todas las celdas
    ws.column_dimensions['A'].width = 40
    for col in range(2, 9):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    # Añadir checkboxes y aplicar estilos a las celdas
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = centered
            cell.border = border
            # Si no es la primera columna (nombre del objeto), añadir checkbox
            if cell.column > 1:
                if isinstance(cell.value, bool):
                    # Usar caracteres Unicode para checkbox
                    cell.value = "☒" if cell.value else "☐"

    # Guardar cambios
    wb.save(output_file)

# Ejemplo de uso
if __name__ == '__main__':
    sample_data = """Accounts	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Ideas	
Checked	Checked	Not Checked	Not Checked	 	 	Not Checked
Account Brands	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Images	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Addresses	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Individuals	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
AI Insight Reasons	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Integration Provider Definitions	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
AiMetadataSyncStatuses	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Internal Organization Units	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
AI Record Insights	
Checked	Not Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Job Profiles	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
App Analytics Query Requests	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Labels	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Appointment Bundle Aggregation Duration Downscales	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Leads	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Appointment Bundle Aggregation Policies	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Locations	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Appointment Bundle Configs	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Location Trust Measures	
Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Appointment Bundle Policies	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Macros	
Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Appointment Bundle Policy Service Territories	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Maintenance Plans	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Appointment Bundle Propagation Policies	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Maintenance Work Rules	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Appointment Bundle Restriction Policies	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Messaging Sessions	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Appointment Bundle Sort Policies	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Messaging Users	
Checked	Checked	Checked	Not Checked	Checked	Not Checked	Not Checked
Appointment Topic Time Slots	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Operating Hours	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Approval Submissions	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Opportunities	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Archive Job Sessions	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Orders	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Assets	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Party Consents	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Asset Account Participants	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Price Books	
Checked	Not Checked	Not Checked	Not Checked	 	 	Not Checked
Asset Contact Participants	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Privacy Consents	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Asynchronous Request Response Events	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Process Exceptions	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Async Operation Trackers	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Products	
Checked	Not Checked	Not Checked	Not Checked	 	 	Not Checked
Attribute Definitions	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Product Items	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Attribute Picklists	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Product Requests	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Attribute Picklist Values	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Product Service Campaigns	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Auth Domain Status Events	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Product Transfers	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Authorization Forms	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Push Topics	
Not Checked	Not Checked	Not Checked	Not Checked	 	 	Not Checked
Authorization Form Consents	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Quick Text	
Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Authorization Form Data Uses	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Rebate Payout Snapshots	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Authorization Form Texts	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Record Action Selectable Item Extracts	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Background Operations	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Record Alerts	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Business Brands	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Record Alert Actionable Targets	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Calculation Procedures	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Recordset Filter Criteria	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Calculation Procedure Steps	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Return Orders	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Calculation Procedure Variables	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Revenue Async Operations	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Calculation Procedure Versions	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Revenue Transaction Error Logs	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Campaigns	
Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Scheduling Constraints	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Cases	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Scheduling Workspaces	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Case Service Processes	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Scheduling Workspace Territories	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Chat Sessions	
Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Scorecards	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Chat Transcripts	
Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Scorecard Associations	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Chat Visitors	
Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Scorecard Metrics	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Communication Subscriptions	
Checked	Not Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Sellers	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Communication Subscription Channel Types	
Checked	Not Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Serialized Products	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Communication Subscription Consents	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Service Appointments	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Communication Subscription Timings	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Service Catalog Item Dependencies	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Contacts	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Service Catalog Requests	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Contact Point Addresses	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Service Catalog Request Related Items	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Contact Point Consents	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Service Contracts	
Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Contact Point Emails	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Service Crews	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Contact Point Phones	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Service Operation Priority Configurations	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Contact Point Type Consents	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Service Resources	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Contact Requests	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Service Resource Preferences	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Context Persistence Events	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Service Territories	
Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Contracts	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Shifts	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Contract Line Items	
Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Shift Patterns	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Contract Line Outcomes	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Shift Scheduling Operations	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Conversation Api Logs	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Shift Templates	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Conversation Api Log Object Summaries	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Shipments	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Customers	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Social Posts	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Data Use Legal Bases	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Solutions	
Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Data Use Purposes	
Checked	Not Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Streaming Channels	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Decision Matrices	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Surveys	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Decision Matrix Columns	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Survey Invitations	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Decision Matrix Column Ranges	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Survey Responses	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Decision Matrix Rows	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Survey Subjects	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Decision Matrix Versions	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Swarms	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Documents	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Swarm Members	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Duplicate Record Sets	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Time Sheets	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Employees	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Travel Modes	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Engagement Attendees	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Units of Measure	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Engagement Channel Types	
Checked	Not Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
User External Credentials	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Engagement Interactions	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Warranty Terms	
Checked	Checked	Checked	Checked	Not Checked	Not Checked	Not Checked
Engagement Topics	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Web Cart Documents	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Entitlements	
Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Work Capacity Availabilities	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Entitlement Contacts	
Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Work Capacity Limits	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Expenses	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Work Capacity Usages	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Expense Reports	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Work Orders	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Expression Sets	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Work Plans	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Expression Set Object Aliases	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Work Plan Selection Rules	
Checked	Checked	Checked	Not Checked	Checked	Not Checked	Not Checked
Expression Set Step Relationships	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Work Plan Templates	
Checked	Checked	Checked	Not Checked	Checked	Not Checked	Not Checked
Expression Set Versions	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Work Step Templates	
Checked	Checked	Checked	Not Checked	Checked	Not Checked	Not Checked
External Managed Accounts	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Work Types	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Geolocation Based Actions	
Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked	Not Checked
Work Type Groups	
Checked	Checked	Checked	Not Checked	Not Checked	Not Checked	Not Checked
Custom Object Permissions"""

    create_permissions_excel(sample_data) 